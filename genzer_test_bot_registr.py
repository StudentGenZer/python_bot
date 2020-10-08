import telebot
import openpyxl
keyboard = telebot.types.ReplyKeyboardMarkup()
keyboard.row("Реєстрація")
key_course = telebot.types.ReplykeyboardMarkup()
key_course.row("Python","Robot","Maths","Unity")
key_vale = telebot.types.ReplyKeyboardMarkup()
key_vale.row("Индив","Группа")
key_anw = telebot.types.ReplyKeyboardMarkup()
key_anw.row("Так","Ні")
bot = telebot.TeleBot('915489974:AAGL39ZFY0YMiGhBelUyApqiET7d7O1z8eU')
log = True
reg = True
log_2 = True
name = ""
number_phone = ""
less_val = ""
@bot.message_handler(commands = ['start'])
def rip(message):
    bot.send_message(message.from_user.id,"Вітаю! До ваших послуг телеграм бот GenerationZ",reply_markup = keyboard)
@bot.message_handler(content_types = ['text'])
def pir(message):
    global log
    global log_2
    if message.text == "Реєстрація":
        wb = openpyxl.load_workbook(filename = 'people.xlsx')
        List = wb["sheet1"]
        
        List['A'+str(len(List)+1)] = message.from_user.id
        msg = bot.send_message(message.from_user.id, "Введіть ім'я")
        bot.register_next_step_handler(msg , step_1)
        log = False
        wb.save('people.xlsx')
def step_1(message):
    global name
    name = message.text
    wb = openpyxl.load_workbook(filename = 'people.xlsx')
    List = wb['sheet1']
    List['B'+str(len(List)+1)] = name
    wb.save('people.xlsx')
    msg = bot.send_message(message.from_user.id, "Введіть номер телефона")
    bot.register_next_step_handler(msg , step_2)
    log = True
    log_2 = False
def step_2(message):
    global phone_number
    number_phone = message.text
    wb = openpyxl.load_workbook(filename = 'people.xlsx')
    List = wb["sheet1"]
    List['C'+str(len(List)+1)] = number_phone
    wb.save('people.xlsx')
    msg = bot.send_message(message.from_user.id, "Виберіть тип занять які ви відвідуєте",reply_markup = key_vale)
    bot.register_next_step_handler(msg , step_3)
def step_3(message):
    global less_val
    if message.text != "Индив" or "Группа":
        msg = bot.send_message(message.from_user.id, "Виберіть тип занять які ви відвідуєте",reply_markup = key_vale)
        bot.register_next_step_handler(msg , step_3)
    else :
        wb = openpyxl.load_workbook(filename = 'people.xlsx')
        List = wb["sheet1"]
        List['D'+str(len(List)+1)] = message.text
        wb.save('people.xlsx')
        msg = bot.send_message(message.from_user.id,"Виберіть курс на який ви ходите",reply_markup = key_course)
        bot.register_next_step_handler(msg, step_4)
def step_4(message):
    if message.text != "Python" or "Robot" or "Maths" or "Unity":
        msg = bot.send_message(message.from_user.id,"Виберіть курс на який ви ходите",reply_markup = key_course)
        bot.register_next_step_handler(msg, step_4)
    else :
        wb = openpyxl.load_workbook(filename = 'people.xlsx')
        List = wb['sheet1']
        List['E'+str(len(List)+1)] = message.text
        wb.save('people.xlsx')
        msg = bot.send_message(message.from_user.id,"Чи відвідуєте ви ще один курс?",reply_markup = key_anw)
        bot.register_next_step_handler(msg, course_1)
def course_1(message):
    if message.text == "Так":
        msg = bot.send_message(message.from_user.id,"Виберіть курс який ви відвідуєте",reply_markup = key_course)
        bot.register_next_step_handler(msg, cousre_2)
    elif message.text == "Ні":
        msg = bot.send_message(message.from_user.id,"Реєстрація завершена")
        bot.register_next_step_handler(msg, end)
def course_2(message):
    if message.text != "Python" or "Robot" or "Maths" or "Unity":
        msg = bot.send_message(message.from_user.id,"Виберіть курс на який ви ходите",reply_markup = key_course)
        bot.register_next_step_handler(msg, course_2)
    else :
        wb = openpyxl.load_workbook(filename = 'people.xlsx')
        List = wb['sheet1']
        List['E'+str(len(List)+1)] = message.text
        wb.save('people.xlsx')
        msg = bot.send_message(message.from_user.id,"Чи відвідуєте ви ще один курс?",reply_markup = key_anw)
        bot.register_next_step_handler(msg, course_3)
def course_3(message):
     if message.text == "Так":
        msg = bot.send_message(message.from_user.id,"Виберіть курс який ви відвідуєте",reply_markup = key_course)
        bot.register_next_step_handler(msg, cousre_4)
    elif message.text == "Ні":
        msg = bot.send_message(message.from_user.id,"Реєстрація завершена")
        bot.register_next_step_handler(msg, end)
def course_4(message):
    wb = openpyxl.load_workbook(filename = 'people.xlsx')
    List = wb["sheet1"]
    List['F'+str(len(List)+1)] = message.text
    wb.save('people.xlsx')
def ends(message):
    wb = openpyxl.load_workbook(filename = 'people.xlsx')
    List = wb["sheet1"]
    List['G'+str(len(List)+1)] = message.text
    wb.save('people.xlsx')
bot.polling()
